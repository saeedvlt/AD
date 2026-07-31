import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import re



MONTH_MAP = {
    "Jan": "01 - Jan",
    "Feb": "02 - Feb",
    "Mar": "03 - Mar",
    "Apr": "04 - Apr",
    "May": "05 - May",
    "Jun": "06 - Jun",
    "Jul": "07 - Jul",
    "Aug": "08 - Aug",
    "Sep": "09 - Sep",
    "Oct": "10 - Oct",
    "Nov": "11 - Nov",
    "Dec": "12 - Dec"
}

META_COLS = {
    1: "Emp #",
    2: "Name",
    3: "Employee Code",
    4: "Date of Hire",
    5: "Site Location",
    6: "AOP File Assignment"
}


def split_name(name):

    if pd.isna(name):
        return "", ""

    name = str(name).strip()

    parts = re.split(
        r'(?<=[a-z])(?=[A-Z])',
        name,
        maxsplit=1
    )

    if len(parts) == 2:
        return parts[0], parts[1]

    return name, ""

def append_manulife(df, wb):
    """
    Read the yellow summary box from the 'Manulife Ins Costs' worksheet
    and append Manu / Insurance records for every employee-month.
    """

    if "Manulife Ins Costs" not in wb.sheetnames:
        return df

    ws = wb["Manulife Ins Costs"]

    # ------------------------------------------------------------------
    # !!! CHANGE THESE CELL REFERENCES TO MATCH YOUR YELLOW BOX !!!
    # ------------------------------------------------------------------

    # Find the header row that contains "Manu"

    header_row = None
    
    for r in range(1, ws.max_row + 1):
        values = [
            str(ws.cell(r, c).value).strip() if ws.cell(r, c).value else ""
            for c in range(1, ws.max_column + 1)
        ]
    
        if "Manu" in values:
            header_row = r
            break
    
    if header_row is None:
        raise ValueError("Could not locate Manu header in 'Manulife Ins Costs'.")
    
    headers = {}
    
    for c in range(1, ws.max_column + 1):
        value = ws.cell(header_row, c).value
    
        if value is not None:
            headers[str(value).strip()] = c
    
    # ---------------------------------------------------------
    # Locate the yellow summary table automatically
    # ---------------------------------------------------------
    
    manu_col = None
    header_row = None
    
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
    
            value = ws.cell(r, c).value
    
            if str(value).strip() == "Manu":
                header_row = r
                manu_col = c
                break
    
        if manu_col:
            break
    
    if manu_col is None:
        raise ValueError("Could not find the 'Manu' header.")
    
    insurance_col = manu_col + 1
    site_col = manu_col - 1
    
    lookup = {}
    
    r = header_row + 1
    
    while True:
    
        site = ws.cell(r, site_col).value
    
        if site in (None, ""):
            break
    
        site = str(site).strip()
    
        lookup[site] = {
            "Manu": ws.cell(r, manu_col).value,
            "Insurance": ws.cell(r, insurance_col).value,
        }
    
        r += 1
        print(lookup)
    # ------------------------------------------------------------------
    # Unique employee/month combinations
    # ------------------------------------------------------------------

    employee_months = df[
        [
            "Emp #",
            "Name",
            "Last Name",
            "First Name",
            "Employee Code",
            "Date of Hire",
            "Site Location",
            "AOP File Assignment",
            "Month",
        ]
    ].drop_duplicates()

    new_rows = []

    for _, emp in employee_months.iterrows():

        site = str(emp["Site Location"]).strip()
        
        if site not in lookup:
            continue

        for benefit, amount in lookup[site].items():

            row = emp.to_dict()
        
            row["Benefit"] = benefit
            row["Amount"] = amount
        
            new_rows.append(row)
    if new_rows:

        df = pd.concat(
            [
                df,
                pd.DataFrame(new_rows),
            ],
            ignore_index=True,
        )

    return df

def convert(uploaded_file):

    wb = load_workbook(
        uploaded_file,
        data_only=True
    )

    if "Benefits" in wb.sheetnames:
    	ws = wb["Benefits"]
    else:
    	raise ValueError("Worksheet 'Benefits' not found.")
    
    benefit_cols = {}

    current_benefit = None

    for idx, col in enumerate(range(7, ws.max_column + 1)):


        top = ws.cell(4, col).value
        low = ws.cell(5, col).value

        if top not in (None, ""):
            current_benefit = str(top).strip()

        if low is None:
            continue

        if isinstance(low, datetime):
            benefit_cols[col] = (
                current_benefit,
                low.strftime("%b")
            )
            continue

        low = str(low).strip()

        if low in (
            "Total",
            "CPP Calc",
            "Amount",
            "Period of Payment",
            ""
        ):
            continue

        if low[:3] in (
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec"
        ):

            benefit_cols[col] = (
                current_benefit,
                low[:3]
            )

    records = []

    for i, row in enumerate(range(7, ws.max_row + 1)):

        employee = {}

        empty = True

        for col, field in META_COLS.items():

            value = ws.cell(row, col).value

            employee[field] = value

            if value not in ("", None):
                empty = False

        if empty:
            continue

        for col, (benefit, month) in benefit_cols.items():

            value = ws.cell(row, col).value

            if value is None:
                value = 0

            rec = employee.copy()

            rec["Benefit"] = benefit
            rec["Month"] = MONTH_MAP.get(month, month)
            rec["Amount"] = value

            records.append(rec)

    df = pd.DataFrame(records)

    df[["Last Name", "First Name"]] = (
        df["Name"]
        .apply(lambda x: pd.Series(split_name(x)))
    )

    cols = list(df.columns)

    cols.remove("Last Name")
    cols.remove("First Name")

    idx = cols.index("Name") + 1

    cols = (
        cols[:idx]
        + ["Last Name", "First Name"]
        + cols[idx:]
    )

    df = df[cols]

    df = append_manulife(df, wb)
    
    return df    
