"""Unpivot base-detail Margin Analysis data and add requested percentages.

Designed for the Budget Database Toolkit. It processes Windsor, Cambridge,
Montreal, and Ithaca and returns a normalized pandas DataFrame.
"""

from __future__ import annotations

from datetime import date, datetime
from numbers import Number
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


LOCATIONS = ("Windsor", "Cambridge", "Montreal", "Ithaca")

TERRITORIES = {
    "Windsor": "CDN",
    "Cambridge": "CDN",
    "Montreal": "CDN",
    "Ithaca": "US",
}

SECTION_NAMES = {
    "sales": "Sales",
    "material": "Material",
    "labour": "Labour",
    "labor": "Labour",
    "overhead": "Overhead",
    "gross margin": "Gross Margin",
}
BUDGET_CATEGORIES = {
    "Die Sets": "Die Sets",
    "Cast Die Sets": "Die Sets",
    "Stght Fwd Die Sets": "Die Sets",

    "Ground Steel": "Plate - Ground / Rough",
    "Rough Steel": "Plate - Ground / Rough",

    "Machined Steel": "Plate - Machined",
    "Bolster Plates": "Plate - Machined",
    "Customer Material": "Plate - Machined",

    "Fabrications": "Fabs",

    "Components": "Components",
}

def clean_text(value: Any) -> str:
    """Return safely trimmed text for a worksheet value."""
    return "" if value is None else str(value).strip()


def budget_category(line_item: str) -> str:
    """Return the budget category for a line item."""
    return BUDGET_CATEGORIES.get(line_item, line_item)

def is_number(value: Any) -> bool:
    """True for spreadsheet numeric values, excluding booleans."""
    return isinstance(value, Number) and not isinstance(value, bool)


def find_month_columns(ws: Worksheet) -> dict[int, datetime]:
    """Find the first contiguous 12-month date series in the top of a sheet."""
    for row in range(1, min(ws.max_row, 15) + 1):
        for start_column in range(1, ws.max_column - 10):
            values = [ws.cell(row, start_column + offset).value for offset in range(12)]

            if not all(isinstance(value, (datetime, date)) for value in values):
                continue

            periods = [
                datetime(value.year, value.month, 1)
                for value in values
            ]

            if [period.month for period in periods] == list(range(1, 13)):
                return {
                    start_column + offset: period
                    for offset, period in enumerate(periods)
                }

    raise ValueError(
        f"Could not find a Jan–Dec date series on worksheet '{ws.title}'."
    )


def is_derived_line(label: str) -> bool:
    """Exclude calculated totals, ratios, and reconciliation lines."""
    normalized = label.casefold()

    return (
        normalized.startswith("total")
        or normalized.startswith("per ")
        or normalized.startswith("sales per")
        or normalized == "diff"
        or normalized.startswith("adjust to actual")
        or normalized.startswith("gross margin")
        or normalized.endswith("%")
        or " per gp" in normalized
        or "financial statement" in normalized
        or normalized in {"f/s", "fabs only", "plug"}
    )


def has_monthly_amount(
    ws: Worksheet,
    row: int,
    month_columns: dict[int, datetime],
) -> bool:
    """Keep a base row only when at least one month contains a value."""
    return any(
        is_number(ws.cell(row, column).value)
        and ws.cell(row, column).value != 0
        for column in month_columns
    )


def percent(
    numerator: pd.Series,
    denominator: pd.Series,
) -> pd.Series:
    """Return percentages on a 0–100 scale."""
    return (
        numerator.div(denominator.where(denominator != 0))
        * 100
    ).round(2)

def add_percentages(data: pd.DataFrame) -> pd.DataFrame:
    """Add all derived percentage calculations."""

    output = data.copy()

    # ---------------------------------------------------------
    # Plant Sales (Sales only)
    # ---------------------------------------------------------
    plant_sales = (
        output.loc[output["Section"] == "Sales"]
        .groupby(
            ["Location", "Period"],
            as_index=False,
        )["Amount"]
        .sum()
        .rename(columns={"Amount": "_plant_sales"})
    )

    output = output.merge(
        plant_sales,
        on=["Location", "Period"],
        how="left",
    )

    # ---------------------------------------------------------
    # Territory totals by category
    # ---------------------------------------------------------
    territory_totals = (
        output.groupby(
            [
                "Territory",
                "Section",
                "Budget Category",
                "Period",
            ],
            as_index=False,
        )["Amount"]
        .sum()
        .rename(columns={"Amount": "_territory_total"})
    )

    output = output.merge(
        territory_totals,
        on=[
            "Territory",
            "Section",
            "Line Item",
            "Period",
        ],
        how="left",
    )

    output["Plant Share of Territory Category %"] = percent(
        output["Amount"],
        output["_territory_total"],
    )

    # ---------------------------------------------------------
    # Category % of Plant Sales
    # ---------------------------------------------------------
    output["Category % of Plant Sales"] = percent(
        output["Amount"],
        output["_plant_sales"],
    )

    # ---------------------------------------------------------
    # Matching Sales lookup
    # ---------------------------------------------------------
    sales_lookup = (
        output.loc[output["Section"] == "Sales"]
        [
            [
                "Location",
                "Line Item",
                "Period",
                "Amount",
            ]
        ]
        .rename(columns={"Amount": "_matching_sales"})
        .drop_duplicates(
            subset=[
                "Location",
                "Line Item",
                "Period",
            ]
        )
    )

    output = output.merge(
        sales_lookup,
        on=[
            "Location",
            "Line Item",
            "Period",
        ],
        how="left",
    )
    annual_category = (
        output.groupby(
            [
                "Territory",
                "Section",
                "Line Item",
            ],
            as_index=False,
        )["Amount"]
        .sum()
        .rename(columns={"Amount": "_annual_category_total"})
    )

    output = output.merge(
        annual_category,
        on=[
            "Territory",
            "Section",
            "Line Item",
        ],
        how="left",
    )

    plant_category = (
        output.groupby(
            [
                "Location",
                "Section",
                "Line Item",
            ],
            as_index=False,
        )["Amount"]
        .sum()
        .rename(columns={"Amount": "_plant_category_total"})
    )
    
    output = output.merge(
        plant_category,
        on=[
            "Location",
            "Section",
            "Line Item",
        ],
        how="left",
    )

    # ---------------------------------------------------------
    # Initialize columns
    # ---------------------------------------------------------
    output["Material % of Matching Sales"] = pd.NA
    output["Labour % of Matching Sales"] = pd.NA
    output["Overhead % of Matching Sales"] = pd.NA
    output["Plant Share of Annual Territory Category %"] = percent(
        output["_plant_category_total"],
        output["_annual_category_total"],
    )

    # ---------------------------------------------------------
    # Material
    # ---------------------------------------------------------
    mask = output["Section"] == "Material"

    output.loc[
        mask,
        "Material % of Matching Sales",
    ] = percent(
        output.loc[mask, "Amount"],
        output.loc[mask, "_matching_sales"],
    )

    # ---------------------------------------------------------
    # Labour
    # ---------------------------------------------------------
    mask = output["Section"] == "Labour"

    output.loc[
        mask,
        "Labour % of Matching Sales",
    ] = percent(
        output.loc[mask, "Amount"],
        output.loc[mask, "_matching_sales"],
    )

    # ---------------------------------------------------------
    # Overhead
    # ---------------------------------------------------------
    mask = output["Section"] == "Overhead"

    output.loc[
        mask,
        "Overhead % of Matching Sales",
    ] = percent(
        output.loc[mask, "Amount"],
        output.loc[mask, "_matching_sales"],
    )

    return output.drop(
        columns=[
            "_plant_sales",
            "_territory_total",
            "_matching_sales",
            "_annual_category_total",
            "_plant_category_total",
        ]
    )
def convert(uploaded_file: Any) -> pd.DataFrame:
    """Return base-detail margin analysis data in long format.

    Derived report rows (totals, ratios, gross margin, differences, and
    ``Adjust to Actual``) are intentionally excluded to prevent double-counting.
    Manual adjustment lines, such as ``Adjust: Rebates``, are retained.
    """

    workbook = load_workbook(uploaded_file, data_only=True)

    missing_sheets = [
        name
        for name in LOCATIONS
        if name not in workbook.sheetnames
    ]

    if missing_sheets:
        raise ValueError(
            "Missing required worksheet(s): "
            + ", ".join(missing_sheets)
        )

    records: list[dict[str, Any]] = []

    for location in LOCATIONS:

        ws = workbook[location]

        month_columns = find_month_columns(ws)

        current_section: str | None = None

        for row in range(1, ws.max_row + 1):

            label = clean_text(ws.cell(row, 1).value)
            normalized_label = label.casefold()

            # Detect section headers
            if normalized_label in SECTION_NAMES:
                current_section = SECTION_NAMES[normalized_label]
                continue

            # Ignore Gross Margin section
            if current_section == "Gross Margin":
                continue

            # Skip blanks and calculated rows
            if (
                not label
                or current_section is None
                or is_derived_line(label)
            ):
                continue

            if not has_monthly_amount(
                ws,
                row,
                month_columns,
            ):
                continue

            # Build normalized records
            for column, period in month_columns.items():

                value = ws.cell(row, column).value

                amount = (
                    round(float(value), 2)
                    if is_number(value)
                    else 0.0
                )

                records.append(
                    {
                        "Location": location,
                        "Territory": TERRITORIES[location],
                        "Section": current_section,
                        "Budget Category": budget_category(label),
                        "Line Item": label,
                        "Period": period,
                        "Month": period.strftime("%m - %b"),
                        "Amount": amount,
                    }
                )
    columns = [
        "Location",
        "Territory",
        "Section",
        "Budget Category",
        "Line Item",
        "Period",
        "Month",
        "Amount",
    ]

    data = pd.DataFrame(records, columns=columns)

    if data.empty:
        return data


    return add_percentages(data)
