import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------
# File paths
# ---------------------------------------------------------
INPUT_FILE = r"NADS, CORP & Outside Sales Allocations Summary 2026.xlsx"
OUTPUT_FILE = "Allocation_Unpivot.xlsx"

MONTHS = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
]

wb = load_workbook(INPUT_FILE, data_only=True)

all_data = []

for sheet in wb.sheetnames:

    if "allocation" not in sheet.lower():
        continue

    ws = wb[sheet]

    # ---------------------------------------------
    # Find the row containing month headers
    # ---------------------------------------------
    month_row = None

    for r in range(1, ws.max_row + 1):

        values = [
            ws.cell(r, c).value
            for c in range(1, ws.max_column + 1)
        ]

        month_count = sum(v in MONTHS for v in values)

        if month_count >= 6:
            month_row = r
            break

    if month_row is None:
        print(f"Skipped {sheet} (no month row found)")
        continue

    # ---------------------------------------------
    # Month columns
    # ---------------------------------------------
    month_cols = {}

    for c in range(1, ws.max_column + 1):

        value = ws.cell(month_row, c).value

        if value in MONTHS:
            month_cols[c] = value

    # ---------------------------------------------
    # Unpivot
    # ---------------------------------------------
    for r in range(month_row + 1, ws.max_row + 1):

        row_values = [
            ws.cell(r, c).value
            for c in range(1, ws.max_column + 1)
        ]

        # Ignore completely blank rows
        if all(v is None for v in row_values):
            continue

        # Find first text cell before month columns
        description = None

        for c in range(1, min(month_cols.keys())):

            value = ws.cell(r, c).value

            if isinstance(value, str) and value.strip():
                description = value.strip()

        if description is None:
            continue

        for col, month in month_cols.items():

            value = ws.cell(r, col).value

            if value is None:
                continue

            all_data.append({
                "Sheet": sheet,
                "Description": description,
                "Month": month,
                "Value": value
            })

# ---------------------------------------------------------
# Export
# ---------------------------------------------------------
df = pd.DataFrame(all_data)

df.to_excel(OUTPUT_FILE, index=False)

print(f"Finished. {len(df):,} rows written to {OUTPUT_FILE}")
