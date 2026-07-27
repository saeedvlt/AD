"""Expense-template converter for the Budget Database Toolkit.

This is a drop-in replacement for ``converters/expense.py``.  It contains no
Streamlit UI code; the toolkit calls ``convert(uploaded_file)``.
"""

from __future__ import annotations

from decimal import Decimal
from numbers import Number
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# These are instruction, summary, or report sheets—not source expense detail.
IGNORE_SHEETS = {
    "Instructions",
    "Totals",
    "Total PP Indirect",
    "By GL Account",
    "Sheet1",
    "Salaries - Head Count Summary",
}

MONTH_MAP = {
    "Jan": "01 - Jan",
    "Feb": "02 - Feb",
    "Mar": "03 - Mar",
    "Apr": "04 - Apr",
    "May": "05 - May",
    "Jun": "06 - Jun",
    "Jul": "07 - Jul",
    "Aug": "08 - Aug",
    "Sep": "09 - Sep",
    "Sept": "09 - Sep",
    "Oct": "10 - Oct",
    "Nov": "11 - Nov",
    "Dec": "12 - Dec",
}

# Header spellings are not fully consistent across the workbook.  These map
# each accepted spelling to the canonical short month name used in MONTH_MAP.
HEADER_MONTHS = {
    "jan": "Jan",
    "january": "Jan",
    "feb": "Feb",
    "february": "Feb",
    "mar": "Mar",
    "march": "Mar",
    "apr": "Apr",
    "april": "Apr",
    "may": "May",
    "jun": "Jun",
    "june": "Jun",
    "jul": "Jul",
    "july": "Jul",
    "aug": "Aug",
    "august": "Aug",
    "sep": "Sep",
    "sept": "Sep",
    "september": "Sep",
    "oct": "Oct",
    "october": "Oct",
    "nov": "Nov",
    "november": "Nov",
    "dec": "Dec",
    "december": "Dec",
}
MONTH_SEQUENCE = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]

# Both spellings occur in the templates.  Matching is case-insensitive.
FUNCTION_NAMES = {
    "operations",
    "cdn sales",
    "hr",
    "finance",
    "sales",
    "engineering",
    "purchasing",
    "quality",
    "it",
    "management",
    "process planning",
    "processing planning",
    "maintenance",
    "production",
}

EXCLUDED_ROW_WORDS = ("subtotal", "total", "head count", "headcount")


def clean_text(value: Any) -> str:
    """Return cell text without turning blank cells into the string 'None'."""
    return "" if value is None else str(value).strip()


def find_header_row(ws: Worksheet) -> int | None:
    """Find the first header row that contains at least six month names."""
    for row in range(1, min(ws.max_row, 40) + 1):
        month_count = sum(
            clean_text(ws.cell(row, column).value).casefold() in HEADER_MONTHS
            for column in range(1, ws.max_column + 1)
        )
        if month_count >= 6:
            return row
    return None


def find_month_columns(ws: Worksheet, header_row: int) -> dict[str, int]:
    """Return Jan–Dec columns identified from their actual header labels.

    Fields such as Start Period, End Period, Amount, and Totals are
    deliberately excluded: only one of the month-name headers is accepted.
    """
    # Some Salary sheets contain additional allocation tables with their own
    # month headings far to the right.  The old dictionary logic overwrote
    # G:R (the budget months) with one of those later tables.  Select the
    # first *contiguous Jan–Dec sequence* instead.
    for start_column in range(1, ws.max_column - len(MONTH_SEQUENCE) + 2):
        candidate = [
            HEADER_MONTHS.get(
                clean_text(ws.cell(header_row, start_column + offset).value).casefold()
            )
            for offset in range(len(MONTH_SEQUENCE))
        ]
        if candidate == MONTH_SEQUENCE:
            return {
                MONTH_MAP[month]: start_column + offset
                for offset, month in enumerate(MONTH_SEQUENCE)
            }

    return {}


def is_summary_or_headcount_row(*values: Any) -> bool:
    """Exclude titles and totals even when their month cells contain formulas."""
    text = " ".join(clean_text(value).casefold() for value in values)
    return any(word in text for word in EXCLUDED_ROW_WORDS)


def is_functional_unit(text: str) -> bool:
    """Recognize functional headings used across the expense templates."""
    normalized = text.casefold()
    return (
        normalized in FUNCTION_NAMES
        or normalized.startswith("process planning -")
        or normalized.startswith("processing planning -")
    )


def as_amount(value: Any) -> float:
    """Return a numeric amount rounded to two decimals; blanks become zero."""
    if value in (None, ""):
        return 0.0
    if isinstance(value, (Number, Decimal)) and not isinstance(value, bool):
        return round(float(value), 2)
    try:
        # Supports numeric values that Excel stored as text, including commas.
        return round(float(str(value).replace(",", "")), 2)
    except (TypeError, ValueError):
        return 0.0


def clean_description(primary_value: Any, fallback_value: Any) -> str | None:
    """Use a text description and never place a rate/FTE number in it."""
    primary = clean_text(primary_value)
    fallback = clean_text(fallback_value)
    if primary:
        try:
            float(primary.replace(",", ""))
        except ValueError:
            return primary
    return fallback or None


def row_has_monthly_data(
    value_ws: Worksheet,
    formula_ws: Worksheet,
    row: int,
    month_columns: dict[str, int],
) -> bool:
    """Return True when a detail row has a monthly value or monthly formula.

    The values workbook gives us calculated amounts.  The formulas workbook
    prevents an Excel formula row from being mistaken for a blank heading when
    its cached value has not yet been refreshed.
    """
    for column in month_columns.values():
        if as_amount(value_ws.cell(row, column).value) != 0:
            return True
        formula = formula_ws.cell(row, column).value
        if isinstance(formula, str) and formula.startswith("="):
            return True
    return False


def convert(uploaded_file: Any) -> pd.DataFrame:
    """Convert an expense workbook into a Power BI-ready DataFrame.

    Behaviour preserved from the original converter:
    - Uses the worksheet name as Expense Category.
    - Carries GL section and functional-unit labels down to detail rows.
    - Outputs every month for a qualifying detail row, including zero months.

    Targeted fixes:
    - Skips Salaries - Head Count Summary entirely.
    - Excludes Total, Sub-total, and Head Count titles/lines.
    - Rounds every Amount to two decimals.
    - Recognizes both Process Planning and Processing Planning.
    """
    # Read calculated values for output and formula cells for reliable row
    # classification.  Streamlit UploadedFile objects need rewinding between
    # the two reads.
    workbook = load_workbook(uploaded_file, data_only=True)
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)
    formula_workbook = load_workbook(uploaded_file, data_only=False)
    records: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        if sheet_name in IGNORE_SHEETS:
            continue

        ws = workbook[sheet_name]
        formula_ws = formula_workbook[sheet_name]
        header_row = find_header_row(ws)
        if header_row is None:
            continue

        month_columns = find_month_columns(ws, header_row)
        if not month_columns:
            continue

        current_gl: str | None = None
        current_function: str | None = None

        for row in range(header_row + 1, ws.max_row + 1):
            column_a = ws.cell(row, 1).value
            column_b = ws.cell(row, 2).value
            column_c = ws.cell(row, 3).value
            text = clean_text(column_a)

            # Column A is the source item's identifier.  Blank-column-A rows
            # are layout/formula rows and must never become output records.
            if not text:
                continue

            # Do this before setting context so Sub-total and Head Count labels
            # cannot overwrite a GL or functional unit for later detail rows.
            if is_summary_or_headcount_row(column_a, column_b, column_c):
                continue

            has_monthly_data = row_has_monthly_data(
                ws,
                formula_ws,
                row,
                month_columns,
            )

            # Functional-unit headings must be checked before GL headings:
            # "Process Planning - DS" contains a dash but is a department,
            # not a GL section.  Its overtime rows inherit this value.
            if is_functional_unit(text):
                current_function = text
                continue

            # A GL section normally has a label such as "Building - 7555"
            # and no monthly amounts.  A data row containing a dash remains a
            # detail row and is not skipped.
            if " - " in text and not has_monthly_data:
                current_gl = text
                current_function = None
                continue

            if not has_monthly_data:
                continue

            item = text or clean_text(column_b) or None

            for month, column in month_columns.items():
                records.append(
                    {
                        "Expense Category": sheet_name,
                        "GL Account": current_gl,
                        "Functional Unit": current_function,
                        "Item": item,
                        "Description": clean_description(column_c, column_b),
                        "Month": month,
                        "Amount": as_amount(ws.cell(row, column).value),
                    }
                )

    columns = [
        "Expense Category",
        "GL Account",
        "Functional Unit",
        "Item",
        "Description",
        "Month",
        "Amount",
    ]
    return pd.DataFrame(records, columns=columns)
