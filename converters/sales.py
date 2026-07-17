
import pandas as pd
from openpyxl import load_workbook


products = {
    7: "Die Sets",
    8: "Plate - Machined",
    9: "Components",
    10: "Fabs",
    11: "Springs",
    12: "Plate - Grnd/Rough",
    15: "Other",
    16: "Total"
}


def convert(uploaded_file):

    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    records = []

    row = 11

    while row <= ws.max_row:

        customer = ws.cell(row, 4).value

        if customer is None:
            row += 1
            continue

        industry = ws.cell(row, 5).value

        metric_rows = [
            row + 1,
            row + 2,
            row + 3,
            row + 4,
            row + 5
        ]

        for r in metric_rows:

            metric = ws.cell(r, 4).value
            year = ws.cell(r, 3).value

            for col, product in products.items():

                value = ws.cell(r, col).value

                records.append({
                    "Customer": customer,
                    "Industry": industry,
                    "Year": year,
                    "Metric": metric,
                    "Product": product,
                    "Sales": value
                })

        row += 7


    df = pd.DataFrame(records)
    return df
