"""Convert an expense budget workbook into a normalized DataFrame.

This module is designed for use by the existing Streamlit toolkit:

    from converters.expense_v2 import convert
    df = convert(uploaded_file)

It intentionally contains no Streamlit UI code.  The original ``expense.py``
can remain in place while this version is tested.
"""

from __future__ import annotations

import re
from collections.abc import Iterable
from numbers import Number
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


IGNORE_SHEETS = {
    "Instructions",
    "Totals",
    "Total PP Indirect",
    "By GL Account",
    "Sheet1",
}

# The final labels sort chronologically in Excel and Power BI.
MONTH_MAP = {
    "jan": "01 - Jan",
    "feb": "02 - Feb",
    "mar": "03 - Mar",
    "apr": "04 - Apr",
    "may": "05 - May",
    "jun": "06 - Jun",
    "jul": "07 - Jul",
    "aug": "08 - Aug",
    "sep": "09 - Sep",
    "sept": "09 - Sep",
    "oct": "10 - Oct",
    "nov": "11 - Nov",
    "dec": "12 - Dec",
}

# These are section labels, rather than expense-line items.
FUNCTION_NAMES = {
    "operations",
    "hr",
    "finance",
    "sales",
    "engineering",
    "purchasing",
    "quality",
    "it",
    "management",
    "processing planning",
    "maintenance",
    "production",
}

# Supports examples such as 7555-912, 7555 - 912, and 7555-912 - Building.
GL_PATTERN = re.compile(
    r"^\s*(?P<code>\d{3,}(?:\s*-\s*\d{2,})+)\s*(?:-\s*(?P<description>.+?))?\s*$"
)

DESCRIPTION_HEADER_NAMES = {
    "description",
    "details",
    "detail",
    "comments",
    "comment",
    "notes",
    "note",
}

NON_DESCRIPTION_HEADERS = {
    "amount",
    "rate",
    "fte",
    "%",
    "percent",
    "percentage",
    "period of payment",
    "payment period",
    "total",
    "cpp calc",
}


def _clean_text(value: Any) -> str:
    """Return a trimmed string, using an empty string for blank cells."""
    return "" if value is None else str(value).strip()


def _month_key(value: Any) -> str | None:
    """Return a normalized month key for month-header text, if it is a month."""
    text = _clean_text(value).lower()
    if not text:
        return None
    return text if text in MONTH_MAP else None


def find_header_row(ws: Worksheet, max_rows_to_check: int = 40) -> int | None:
    """Find the first row containing at least six recognizable month headers."""
    for row_number in range(1, min(ws.max_row, max_rows_to_check) + 1):
        month_count = sum(
            _month_key(ws.cell(row_number, column).value) is not None
            for column in range(1, ws.max_column + 1)
        )
        if month_count >= 6:
            return row_number
    return None


def find_month_columns(ws: Worksheet, header_row: int) -> dict[str, int]:
    """Map final month labels (``01 - Jan``) to their worksheet columns."""
    columns: dict[str, int] = {}
    for column in range(1, ws.max_column + 1):
        key = _month_key(ws.cell(header_row, column).value)
        if key:
            columns[MONTH_MAP[key]] = column
    return columns


def _header_texts(ws: Worksheet, header_row: int, column: int) -> Iterable[str]:
    """Yield useful header labels from the month-header row and two rows above it."""
    for row in range(max(1, header_row - 2), header_row + 1):
        text = _clean_text(ws.cell(row, column).value).lower()
        if text:
            yield text


def find_description_column(
    ws: Worksheet, header_row: int, month_columns: dict[str, int]
) -> int | None:
    """Find an explicitly labelled text-description column.

    Only an actual description-like header is accepted.  This deliberately
    avoids the old fallback of treating column C as a description field, which
    incorrectly captured rate/FTE/percentage values on some worksheets.
    """
    month_column_numbers = set(month_columns.values())

    for column in range(1, ws.max_column + 1):
        if column in month_column_numbers:
            continue

        labels = set(_header_texts(ws, header_row, column))
        if labels & DESCRIPTION_HEADER_NAMES and not labels & NON_DESCRIPTION_HEADERS:
            return column
    return None


def find_item_column(ws: Worksheet, header_row: int, month_columns: dict[str, int]) -> int:
    """Choose the item column from a header when possible, otherwise use column A.

    The template normally stores hierarchy labels and item names in column A.
    A recognised item-style header can override this without relying on a fixed
    description column.
    """
    preferred_headers = {
        "item",
        "vendor",
        "service provider",
        "employee",
        "employee name",
        "name",
        "account",
        "account description",
    }
    month_column_numbers = set(month_columns.values())

    for column in range(1, ws.max_column + 1):
        if column in month_column_numbers:
            continue
        if set(_header_texts(ws, header_row, column)) & preferred_headers:
            return column
    return 1


def parse_gl_header(text: str) -> tuple[str | None, str | None] | None:
    """Return ``(GL Code, GL Description)`` when *text* is a GL section row."""
    match = GL_PATTERN.match(text)
    if not match:
        return None

    code = re.sub(r"\s*-\s*", "-", match.group("code"))
    description = _clean_text(match.group("description")) or None
    return code, description


def is_functional_unit(text: str) -> bool:
    """Identify known functional-unit headings without case sensitivity."""
    return text.casefold() in FUNCTION_NAMES


def _is_real_description(value: Any) -> str | None:
    """Keep text descriptions only; discard numeric rate/FTE/percentage values."""
    if value is None or isinstance(value, Number):
        return None

    text = _clean_text(value)
    if not text:
        return None

    # Numeric strings (e.g. "0.04") should not be presented as descriptions.
    try:
        float(text.replace(",", ""))
        return None
    except ValueError:
        return text


def _has_monthly_value(ws: Worksheet, row: int, month_columns: dict[str, int]) -> bool:
    """Return True when at least one month has a non-zero numeric value."""
    for column in month_columns.values():
        value = ws.cell(row, column).value
        if isinstance(value, Number) and value != 0:
            return True
    return False


def extract_sheet_records(ws: Worksheet, sheet_name: str) -> list[dict[str, Any]]:
    """Extract normalized records from one expense worksheet."""
    header_row = find_header_row(ws)
    if header_row is None:
        return []

    month_columns = find_month_columns(ws, header_row)
    if not month_columns:
        return []

    item_column = find_item_column(ws, header_row, month_columns)
    description_column = find_description_column(ws, header_row, month_columns)

    records: list[dict[str, Any]] = []
    current_gl_code: str | None = None
    current_gl_description: str | None = None
    current_function: str | None = None

    for row in range(header_row + 1, ws.max_row + 1):
        item_text = _clean_text(ws.cell(row, item_column).value)

        # GL and functional-unit rows establish context for the detail rows below.
        gl_parts = parse_gl_header(item_text)
        if gl_parts:
            current_gl_code, current_gl_description = gl_parts
            current_function = None
            continue

        if is_functional_unit(item_text):
            current_function = item_text
            continue

        # Preserve existing behaviour: do not output completely zero/blank lines.
        if not _has_monthly_value(ws, row, month_columns):
            continue

        # Retain a useful fallback item if a template uses column B for names.
        item = item_text or _clean_text(ws.cell(row, 2).value) or None
        description = (
            _is_real_description(ws.cell(row, description_column).value)
            if description_column is not None
            else None
        )

        for month, column in month_columns.items():
            amount = ws.cell(row, column).value
            if amount is None:
                amount = 0

            records.append(
                {
                    "Expense Category": sheet_name,
                    "GL Code": current_gl_code,
                    "GL Description": current_gl_description,
                    "Functional Unit": current_function,
                    "Item": item,
                    "Description": description,
                    "Month": month,
                    "Amount": amount,
                }
            )

    return records


def convert(uploaded_file: Any) -> pd.DataFrame:
    """Convert an uploaded expense workbook to a Power BI-ready DataFrame.

    Parameters
    ----------
    uploaded_file:
        A Streamlit UploadedFile, file path, or file-like object accepted by
        ``openpyxl.load_workbook``.
    """
    workbook = load_workbook(uploaded_file, data_only=True)
    records: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        if sheet_name in IGNORE_SHEETS:
            continue
        records.extend(extract_sheet_records(workbook[sheet_name], sheet_name))

    columns = [
        "Expense Category",
        "GL Code",
        "GL Description",
        "Functional Unit",
        "Item",
        "Description",
        "Month",
        "Amount",
    ]
    return pd.DataFrame(records, columns=columns)
