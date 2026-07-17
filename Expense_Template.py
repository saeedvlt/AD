import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(
    page_title="Expense Template Converter",
    page_icon="📊",
    layout="wide"
)

st.title("📊 Expense Template Converter")
st.write(
    "Upload an expense budget workbook and convert every worksheet into one Power BI-ready database."
)

uploaded_file = st.file_uploader(
    "Upload Excel Workbook",
    type=["xlsx"]
)

IGNORE_SHEETS = {
    "Instructions",
    "Totals",
    "Total PP Indirect",
    "By GL Account",
    "Sheet1"
}

MONTHS = [
    "Jan","Feb","Mar","Apr","May","Jun",
    "Jul","Aug","Sep","Sept","Oct","Nov","Dec"
]

FUNCTION_NAMES = {
    "Operations",
    "HR",
    "Finance",
    "Sales",
    "Engineering",
    "Purchasing",
    "Quality",
    "IT",
    "Management",
    "Processing Planning",
    "Maintenance",
    "Production"
}


def find_header_row(ws):

    for r in range(1, min(ws.max_row, 40)+1):

        values = []

        for c in range(1, ws.max_column+1):

            v = ws.cell(r,c).value

            values.append("" if v is None else str(v).strip())

        if sum(v in MONTHS for v in values) >= 6:
            return r

    return None


def month_columns(ws, header):

    cols = {}

    for c in range(1, ws.max_column+1):

        value = ws.cell(header,c).value

        if value is None:
            continue

        value = str(value).strip()

        if value in MONTHS:
            cols[value[:3]] = c

    return cols


if uploaded_file:

    wb = load_workbook(uploaded_file, data_only=True)

    records = []

    progress = st.progress(0)

    sheets = [
        s for s in wb.sheetnames
        if s not in IGNORE_SHEETS
    ]

    for i, sheet in enumerate(sheets):

        progress.progress((i+1)/len(sheets))

        ws = wb[sheet]

        header = find_header_row(ws)

        if header is None:
            continue

        months = month_columns(ws, header)

        current_gl = None
        current_function = None

        for r in range(header+1, ws.max_row+1):

            a = ws.cell(r,1).value
            b = ws.cell(r,2).value
            c = ws.cell(r,3).value

            text = "" if a is None else str(a).strip()

            if " - " in text:

                current_gl = text
                current_function = None
                continue

            if text in FUNCTION_NAMES:

                current_function = text
                continue

            values = {}

            has_data = False

            for month,col in months.items():

                value = ws.cell(r,col).value

                if value is None:
                    value = 0

                values[month] = value

                if isinstance(value,(int,float)) and value != 0:
                    has_data = True

            if not has_data:
                continue

            item = text if text else b

            for month,value in values.items():

                records.append({

                    "Expense Category": sheet,
                    "GL Account": current_gl,
                    "Functional Unit": current_function,
                    "Item": item,
                    "Description": c,
                    "Month": month,
                    "Amount": value

                })

    df = pd.DataFrame(records)

    st.success(f"Finished! {len(df):,} records created.")

    st.dataframe(df.head(100), use_container_width=True)

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

    output.seek(0)

    st.download_button(
        "📥 Download Database",
        data=output,
        file_name="Expense_Database.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
