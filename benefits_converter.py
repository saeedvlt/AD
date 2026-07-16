import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from io import BytesIO
import re

st.set_page_config(
    page_title="Benefits Template Converter",
    page_icon="👥",
    layout="wide"
)

st.title("👥 Benefits Template Converter")

st.write(
    "Upload a Benefits Budget workbook and convert it into a Power BI-ready database."
)

uploaded_file = st.file_uploader(
    "Upload Excel Workbook",
    type=["xlsx"]
)

MONTH_MAP = {
    "Jan": "01 - Jan",
    "Feb": "02 - Feb",
    "Mar": "03 - Mar",
    "Apr": "04 - Apr",
    "May": "05 - May",
    "Jun": "06 - Jun",
    "Jul": "07 - Jul",
    "Aug": "08 - Aug",
    "Sep": "09 - Sep",
    "Oct": "10 - Oct",
    "Nov": "11 - Nov",
    "Dec": "12 - Dec"
}

META_COLS = {
    1: "Emp #",
    2: "Name",
    3: "Employee Code",
    4: "Date of Hire",
    5: "Site Location",
    6: "AOP File Assignment"
}


def split_name(name):

    if pd.isna(name):
        return "", ""

    name = str(name).strip()

    parts = re.split(
        r'(?<=[a-z])(?=[A-Z])',
        name,
        maxsplit=1
    )

    if len(parts) == 2:
        return parts[0], parts[1]

    return name, ""


if uploaded_file:

    wb = load_workbook(
        uploaded_file,
        data_only=True
    )

    if "Benefits" in wb.sheetnames:
        ws = wb["Benefits"]
    else:
        st.error("Worksheet 'Benefits' not found.")
        st.stop()

    benefit_cols = {}

    current_benefit = None

    progress = st.progress(0)

    total_cols = ws.max_column - 6

    for idx, col in enumerate(range(7, ws.max_column + 1)):

        progress.progress(idx / total_cols)

        top = ws.cell(4, col).value
        low = ws.cell(5, col).value

        if top not in (None, ""):
            current_benefit = str(top).strip()

        if low is None:
            continue

        if isinstance(low, datetime):
            benefit_cols[col] = (
                current_benefit,
                low.strftime("%b")
            )
            continue

        low = str(low).strip()

        if low in (
            "Total",
            "CPP Calc",
            "Amount",
            "Period of Payment",
            ""
        ):
            continue

        if low[:3] in (
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec"
        ):

            benefit_cols[col] = (
                current_benefit,
                low[:3]
            )

    records = []

    total_rows = ws.max_row - 6

    for i, row in enumerate(range(7, ws.max_row + 1)):

        progress.progress(i / total_rows)

        employee = {}

        empty = True

        for col, field in META_COLS.items():

            value = ws.cell(row, col).value

            employee[field] = value

            if value not in ("", None):
                empty = False

        if empty:
            continue

        for col, (benefit, month) in benefit_cols.items():

            value = ws.cell(row, col).value

            if value is None:
                value = 0

            rec = employee.copy()

            rec["Benefit"] = benefit
            rec["Month"] = MONTH_MAP.get(month, month)
            rec["Amount"] = value

            records.append(rec)

    df = pd.DataFrame(records)

    df[["Last Name", "First Name"]] = (
        df["Name"]
        .apply(lambda x: pd.Series(split_name(x)))
    )

    cols = list(df.columns)

    cols.remove("Last Name")
    cols.remove("First Name")

    idx = cols.index("Name") + 1

    cols = (
        cols[:idx]
        + ["Last Name", "First Name"]
        + cols[idx:]
    )

    df = df[cols]

    st.success(f"Created {len(df):,} records.")

    st.dataframe(
        df.head(100),
        use_container_width=True
    )

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            index=False
        )

    output.seek(0)

    st.download_button(
        "📥 Download Database",
        data=output,
        file_name="Benefits_Database.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
